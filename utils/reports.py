import os
import sqlite3
from datetime import datetime

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton


def is_admin(user_id):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    c.execute("SELECT * FROM admins WHERE tg_id = ?", (user_id,))
    admin_entry = c.fetchone()
    conn.close()
    return admin_entry is not None


def register_handlers(bot):
    def report(message):
        markup = InlineKeyboardMarkup()

        # Create buttons for different subscription periods
        active = InlineKeyboardButton("Подписки активные на данный момент", callback_data="active")
        expired = InlineKeyboardButton("Истекшие Подписки (за прошлый месяц)", callback_data="expired")
        expired_old = InlineKeyboardButton("Истекшие Подписки (больше месяца назад)", callback_data="expired_old")
        markup.add(active)
        markup.add(expired)
        markup.add(expired_old)
        bot.send_message(message.chat.id, "Какой из отчетов о подписках необходимо сгенерировать?", reply_markup=markup,
                         parse_mode='Markdown')

    def get_username_from_user_id(user_id):
        try:
            # Use the bot's get_chat method to retrieve username
            chat = bot.get_chat(user_id)
            return chat.username if chat.username else '-'
        except Exception as e:
            return '-'

    # Function to fetch all subscribers and export to an Excel file
    def export_subscribers_to_excel():
        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        c.execute("SELECT user_id, phone_number, subscription_date, expiry_date FROM subscribers")
        subscribers = c.fetchall()
        conn.close()

        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Subscribers"

        # Set up header row
        headers = ["User ID", "Username", "Номера Телефона", "Дата начала Подписки", "Дата окончания Подписки"]
        sheet.append(headers)

        # Create a date style for proper date formatting
        date_style = NamedStyle(name="datetime", number_format="DD-MM-YY")

        # Add data rows
        for row in subscribers:
            # Format subscription and expiry date to dd-mm-yy
            subscription_date = datetime.fromisoformat(row[2]).strftime("%d-%m-%y")
            expiry_date = datetime.fromisoformat(row[3]).strftime("%d-%m-%y")

            # Retrieve the username from user_id
            username = get_username_from_user_id(row[0])

            # Add row, converting User ID to string to avoid scientific notation
            sheet.append([str(row[0]), username, row[1], subscription_date, expiry_date])

        # Apply date style to date columns
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
            for cell in row:
                cell.style = date_style

        # Adjust column widths to fit the content
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Find the maximum length of the column content
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Set specific column widths
        sheet.column_dimensions['A'].width = 15  # Column A for User ID
        sheet.column_dimensions['B'].width = 20  # Column B for Username

        # Save the Excel file
        filename = 'Отчет.xlsx'
        workbook.save(filename)

        return filename

    def export_expired_users_to_excel():
        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        c.execute("SELECT user_id, phone_number, subscription_date, expiry_date, expired_date FROM expired")
        expired_users = c.fetchall()
        conn.close()

        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Expired Users"

        # Set up header row
        headers = ["User ID", "Username", "Номера Телефона", "Дата начала Подписки", "Дата окончания Подписки"]
        sheet.append(headers)

        # Create a date style for proper date formatting
        date_style = NamedStyle(name="datetime", number_format="DD-MM-YY")

        # Add data rows
        for row in expired_users:
            subscription_date = datetime.fromisoformat(row[2]).strftime("%d-%m-%y")
            expiry_date = datetime.fromisoformat(row[3]).strftime("%d-%m-%y")

            # Retrieve the username from user_id
            username = get_username_from_user_id(row[0])

            # Add row, converting User ID to string to avoid scientific notation
            sheet.append([str(row[0]), username, row[1], subscription_date, expiry_date])

        # Apply date style to date columns
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=6):
            for cell in row:
                cell.style = date_style

        # Adjust column widths to fit the content
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Find the maximum length of the column content
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Set specific column widths
        sheet.column_dimensions['A'].width = 15  # Column A for User ID
        sheet.column_dimensions['B'].width = 20  # Column B for Username

        # Save the Excel file
        filename = 'Отчет.xlsx'
        workbook.save(filename)

        return filename

    def export_old_expired_users_to_excel():
        conn = sqlite3.connect('expired.db')
        c = conn.cursor()
        c.execute("SELECT user_id, phone_number, subscription_date, expiry_date FROM old_expired")
        old_expired_users = c.fetchall()
        conn.close()

        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Old Expired Users"

        # Set up header row
        headers = ["User ID", "Username", "Номера Телефона", "Дата начала Подписки", "Дата окончания Подписки"]
        sheet.append(headers)

        # Create a date style for proper date formatting
        date_style = NamedStyle(name="datetime", number_format="DD-MM-YY")

        # Add data rows
        for row in old_expired_users:
            subscription_date = datetime.fromisoformat(row[2]).strftime("%d-%m-%y")
            expiry_date = datetime.fromisoformat(row[3]).strftime("%d-%m-%y")

            # Retrieve the username from user_id
            username = get_username_from_user_id(row[0])

            # Add row, converting User ID to string to avoid scientific notation
            sheet.append([str(row[0]), username, row[1], subscription_date, expiry_date])

        # Apply date style to date columns
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=6):
            for cell in row:
                cell.style = date_style

        # Adjust column widths to fit the content
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Find the maximum length of the column content
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Set specific column widths
        sheet.column_dimensions['A'].width = 15  # Column A for User ID
        sheet.column_dimensions['B'].width = 20  # Column B for Username

        # Save the Excel file
        filename = 'Отчет.xlsx'
        workbook.save(filename)

        return filename

    # Command to send expired users report
    def send_expired_users_excel(call):
        if is_admin(call.message.chat.id):
            try:
                # Export expired users to Excel
                file_path = export_expired_users_to_excel()

                # Open the file and send it via the bot
                with open(file_path, 'rb') as file:
                    bot.send_message(call.message.chat.id, "Вот отчет о пользователях с истекшей подпиской:")
                    bot.send_document(call.message.chat.id, file)

                # Optionally delete the file after sending
                os.remove(file_path)

            except Exception as e:
                bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")
        else:
            bot.send_message(call.message.chat.id, "У вас недостаточно прав.")

    def send_old_expired_users_excel(call):
        if is_admin(call.message.chat.id):
            try:
                # Export old expired users to Excel
                file_path = export_old_expired_users_to_excel()

                # Open the file and send it via the bot
                with open(file_path, 'rb') as file:
                    bot.send_message(call.message.chat.id,
                                     "Вот отчет о пользователях, чья подписка истекла более месяца назад:")
                    bot.send_document(call.message.chat.id, file)

                # Optionally delete the file after sending
                os.remove(file_path)

            except Exception as e:
                bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")
        else:
            bot.send_message(call.message.chat.id, "У вас недостаточно прав.")

    def send_subscribers_excel(call):
        if is_admin(call.message.chat.id):
            try:
                # Export data to Excel
                file_path = export_subscribers_to_excel()

                # Open the file and send it via the bot
                with open(file_path, 'rb') as file:
                    bot.send_message(call.message.chat.id, "Вот отчет о всех действующих подписках на КөмекБай:")
                    bot.send_document(call.message.chat.id, file)

                # Optionally delete the file after sending
                os.remove(file_path)

            except Exception as e:
                bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")
        else:
            bot.send_message(call.message.chat.id, "У вас недостаточно прав.")

    bot.register_callback_query_handler(send_expired_users_excel, lambda call: call.data == 'expired')
    bot.register_callback_query_handler(send_old_expired_users_excel, lambda call: call.data == 'expired_old')
    bot.register_callback_query_handler(send_subscribers_excel, lambda call: call.data == 'active')
    bot.register_message_handler(report, commands=['report'])
